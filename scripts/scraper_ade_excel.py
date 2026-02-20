#!/usr/bin/env python3
"""
Scraper XLSX per codici tributo F24 - Agenzia delle Entrate

Scarica e parse file Excel ufficiali da:
https://www.agenziaentrate.gov.it/portale/strumenti/codici-attivita-e-tributo/
f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24

Usage:
    python scraper_ade_excel.py [--update-db] [--export-csv] [--export-json] [--verbose]
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import json
import csv
from datetime import datetime
from pathlib import Path
import sys
import os
from io import BytesIO
from typing import List, Dict, Optional
import argparse


class ScraperADEExcel:
    """Scraper per file Excel ufficiali ADE con codici tributo F24"""
    
    BASE_URL = "https://www.agenziaentrate.gov.it"
    
    # URL pagine con link ai file Excel
    URL_ERARIO = f"{BASE_URL}/portale/strumenti/codici-attivita-e-tributo/f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24/tabella-codici-tributo-erariali-e-regionali"
    URL_TRIBUTI_LOCALI = f"{BASE_URL}/portale/strumenti/codici-attivita-e-tributo/f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24/tabelle-codici-per-tributi-locali"
    URL_INPS = f"{BASE_URL}/portale/strumenti/codici-attivita-e-tributo/f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24/tabelle-codici-inps-e-enti-previdenziali-ed-assicurativi"
    
    def __init__(self, verbose=False):
        self.verbose = verbose
        self.session = requests.Session()
        self.session.headers.update({
            'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })
        self.codici = []
        
    def log(self, msg):
        """Log con timestamp se verbose"""
        if self.verbose:
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"[{timestamp}] {msg}")
    
    def fetch_page(self, url: str) -> Optional[BeautifulSoup]:
        """Scarica pagina HTML"""
        try:
            self.log(f"Scaricamento pagina: {url}")
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            return BeautifulSoup(response.content, 'html.parser')
        except Exception as e:
            print(f"❌ Errore scaricamento {url}: {e}")
            return None
    
    def download_excel(self, url: str) -> Optional[openpyxl.Workbook]:
        """Scarica file Excel da URL"""
        try:
            self.log(f"Download Excel: {url}")
            response = self.session.get(url, timeout=30)
            response.raise_for_status()
            
            # Carica Excel da bytes
            excel_file = BytesIO(response.content)
            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            self.log(f"✓ File Excel caricato: {len(wb.worksheets)} fogli")
            return wb
        except Exception as e:
            print(f"❌ Errore download Excel {url}: {e}")
            return None
    
    def parse_excel_erariali(self, wb: openpyxl.Workbook, sezione: str = "erario") -> List[Dict]:
        """Parse file Excel codici erariali/regionali"""
        codici = []
        
        try:
            sheet = wb.active
            self.log(f"Parse foglio: {sheet.title}, righe: {sheet.max_row}")
            
            # Trova header row (di solito prima o seconda riga)
            header_row = None
            for row_idx in range(1, min(5, sheet.max_row + 1)):
                row_values = [cell.value for cell in sheet[row_idx]]
                # Cerca colonne tipiche: "codice", "descrizione"
                if any(val and isinstance(val, str) and 'codice' in val.lower() for val in row_values):
                    header_row = row_idx
                    self.log(f"Header trovato a riga {row_idx}")
                    break
            
            if not header_row:
                self.log("⚠️  Header non trovato, uso riga 1")
                header_row = 1
            
            # Leggi header
            headers = []
            for cell in sheet[header_row]:
                val = str(cell.value).strip() if cell.value else ""
                headers.append(val.lower())
            
            self.log(f"Colonne: {headers}")
            
            # Trova indici colonne importanti
            idx_codice = self._find_column_index(headers, ['codice tributo', 'codice', 'cod.'])
            idx_desc = self._find_column_index(headers, ['descrizione', 'causale contributo'])
            
            if idx_codice is None:
                print(f"❌ Colonna 'codice' non trovata in: {headers}")
                return []
            
            self.log(f"Colonne identificate - codice: {idx_codice}, descrizione: {idx_desc}")
            
            # Parse righe
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                row = sheet[row_idx]
                
                # Leggi codice
                codice = row[idx_codice].value if idx_codice < len(row) else None
                if not codice:
                    continue
                    
                codice = str(codice).strip()
                
                # Skip righe vuote o intestazioni
                if not codice or len(codice) < 2 or codice.lower() in ['codice', 'cod.']:
                    continue
                
                # Leggi descrizione
                descrizione = ""
                if idx_desc is not None and idx_desc < len(row):
                    desc_val = row[idx_desc].value
                    descrizione = str(desc_val).strip() if desc_val else ""
                
                # Skip se descrizione troppo corta
                if len(descrizione) < 5:
                    continue
                
                # Identifica sotto-sezione
                sotto_sezione = self._identifica_sezione(codice, descrizione)
                
                # Rileva se obsoleto
                attivo = not self._e_obsoleto(descrizione)
                
                codice_dict = {
                    'codice': codice,
                    'sezione': sotto_sezione,
                    'descrizione': descrizione,
                    'causale': self._estrai_causale(descrizione),
                    'periodicita': self._estrai_periodicita(descrizione),
                    'attivo': attivo,
                    'fonte': 'ADE',
                    'data_aggiornamento': datetime.now().date().isoformat()
                }
                
                codici.append(codice_dict)
                
                if self.verbose and len(codici) % 50 == 0:
                    self.log(f"Processati {len(codici)} codici...")
            
            self.log(f"✓ Parse completato: {len(codici)} codici")
            
        except Exception as e:
            print(f"❌ Errore parse Excel: {e}")
            import traceback
            traceback.print_exc()
        
        return codici
    
    def parse_excel_tributi_locali(self, wb: openpyxl.Workbook) -> List[Dict]:
        """Parse file Excel tributi locali (IMU, TASI, TARI) - struttura specifica"""
        codici = []
        
        try:
            sheet = wb.active
            self.log(f"Parse foglio tributi locali: {sheet.title}, righe: {sheet.max_row}")
            
            # File tributi locali hanno struttura diversa
            # Cerca header row
            header_row = None
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                row_values = [str(cell.value).lower() if cell.value else "" for cell in sheet[row_idx]]
                # Cerca "codice ente" o "codice tributo" o "imposta"
                if any('codice' in val or 'ente' in val or 'tributo' in val or 'imposta' in val for val in row_values):
                    header_row = row_idx
                    self.log(f"Header tributi locali trovato a riga {row_idx}")
                    break
            
            if not header_row:
                self.log("⚠️  Header tributi locali non trovato, skip")
                return []
            
            # Leggi header
            headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[header_row]]
            self.log(f"Colonne tributi locali: {headers[:10]}...")  # Prime 10
            
            # Trova indici colonne
            idx_codice_ente = self._find_column_index(headers, ['codice ente', 'codice catastale', 'cod. ente'])
            idx_ente = self._find_column_index(headers, ['ente', 'comune', 'denominazione'])
            idx_cod_tributo = self._find_column_index(headers, ['codice tributo', 'cod. tributo', 'tributo'])
            idx_imposta = self._find_column_index(headers, ['imposta', 'tipo imposta', 'tipologia'])
            
            if idx_cod_tributo is None and idx_imposta is None:
                self.log(f"❌ Colonne tributo/imposta non trovate")
                return []
            
            self.log(f"Colonne identificate - ente: {idx_codice_ente}, tributo: {idx_cod_tributo}, imposta: {idx_imposta}")
            
            # Parse righe
            for row_idx in range(header_row + 1, sheet.max_row + 1):
                row = sheet[row_idx]
                
                # Leggi codice ente (codice catastale comune)
                codice_ente = str(row[idx_codice_ente].value).strip() if idx_codice_ente is not None and idx_codice_ente < len(row) and row[idx_codice_ente].value else None
                
                # Leggi ente/comune
                ente = str(row[idx_ente].value).strip() if idx_ente is not None and idx_ente < len(row) and row[idx_ente].value else ""
                
                # Leggi codice tributo
                cod_tributo = None
                if idx_cod_tributo is not None and idx_cod_tributo < len(row):
                    cod_tributo = row[idx_cod_tributo].value
                    if cod_tributo:
                        cod_tributo = str(cod_tributo).strip()
                
                # Leggi imposta
                imposta = None
                if idx_imposta is not None and idx_imposta < len(row):
                    imposta = row[idx_imposta].value
                    if imposta:
                        imposta = str(imposta).strip()
                
                # Il "codice" per tributi locali è concatenazione: codice_tributo o imposta
                codice_completo = cod_tributo or imposta
                
                if not codice_completo or len(codice_completo) < 2:
                    continue
                
                # Costruisci descrizione
                if ente:
                    descrizione = f"{codice_completo} - {ente}"
                    if imposta and cod_tributo:
                        descrizione = f"{cod_tributo} - {imposta} - {ente}"
                else:
                    descrizione = codice_completo
                
                # Skip se troppo corto
                if len(descrizione) < 5:
                    continue
                
                codice_dict = {
                    'codice': codice_completo,
                    'sezione': 'imu',  # Tutti tributi locali
                    'descrizione': descrizione,
                    'causale': self._estrai_causale(descrizione),
                    'periodicita': 'Annuale',  # IMU/TARI di solito annuali
                    'attivo': True,  # Assume attivi
                    'fonte': 'ADE',
                    'data_aggiornamento': datetime.now().date().isoformat()
                }
                
                codici.append(codice_dict)
                
                if self.verbose and len(codici) % 500 == 0:
                    self.log(f"Processati {len(codici)} codici tributi locali...")
            
            self.log(f"✓ Parse tributi locali completato: {len(codici)} codici")
            
        except Exception as e:
            print(f"❌ Errore parse Excel tributi locali: {e}")
            import traceback
            traceback.print_exc()
        
        return codici
    
    def _find_column_index(self, headers: List[str], keywords: List[str]) -> Optional[int]:
        """Trova indice colonna da keywords"""
        for keyword in keywords:
            for idx, header in enumerate(headers):
                if keyword.lower() in header.lower():
                    return idx
        return None
    
    def _identifica_sezione(self, codice: str, descrizione: str) -> str:
        """Identifica sezione del codice"""
        desc_lower = descrizione.lower()
        
        # IMU/TASI/TARI (tributi locali)
        if any(kw in desc_lower for kw in ['imu', 'tasi', 'tari', 'tributi locali', 'comunale']):
            return 'imu'
        
        # INPS
        if any(kw in desc_lower for kw in ['inps', 'contributi', 'previdenziali', 'gestione separata']):
            return 'inps'
        
        # INAIL
        if 'inail' in desc_lower:
            return 'inail'
        
        # Regionali
        if any(kw in desc_lower for kw in ['regional', 'irap', 'addizionale regionale']):
            return 'regioni'
        
        # Accise
        if 'accise' in desc_lower or 'energetici' in desc_lower:
            return 'accise'
        
        # Default: erario
        return 'erario'
    
    def _e_obsoleto(self, descrizione: str) -> bool:
        """Rileva se codice è obsoleto"""
        parole_obsoleto = [
            'soppresso', 'abolito', 'abrogato', 'cessato',
            'dismesso', 'non più', 'ex ', 'fino al'
        ]
        desc_lower = descrizione.lower()
        return any(parola in desc_lower for parola in parole_obsoleto)
    
    def _estrai_causale(self, descrizione: str) -> str:
        """Estrae causale breve da descrizione"""
        # Prendi prime 100 caratteri
        causale = descrizione[:100]
        # Rimuovi caratteri strani
        causale = re.sub(r'[^\w\s\-\.,()]', '', causale)
        return causale.strip()
    
    def _estrai_periodicita(self, descrizione: str) -> str:
        """Estrae periodicità da descrizione"""
        desc_lower = descrizione.lower()
        
        if 'trimestral' in desc_lower:
            return 'Trimestrale'
        if 'mensil' in desc_lower:
            return 'Mensile'
        if 'annual' in desc_lower or 'annua' in desc_lower:
            return 'Annuale'
        if 'unico' in desc_lower or 'una tantum' in desc_lower:
            return 'Unico'
        
        return 'Variabile'
    
    def scrape_erario_regioni(self) -> List[Dict]:
        """Scarica codici ERARIO e REGIONI da file Excel"""
        print("\n🏛️  Scaricamento codici ERARIO e REGIONI...")
        
        # URL diretti ai file Excel (aggiornati 11/11/2025)
        # Fonte: https://www.agenziaentrate.gov.it/portale/strumenti/codici-attivita-e-tributo/
        #        f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24/
        #        tabella-codici-tributo-erariali-e-regionali
        
        excel_urls = [
            # Tabella ordinata per codice tributo (file principale)
            f"{self.BASE_URL}/portale/documents/20143/448357/TABELLA_EREL_11_11_2025_codice+tributo.xlsx/7195eadf-0fe4-7662-a391-59965b747eb8"
        ]
        
        codici_totali = []
        
        for excel_url in excel_urls:
            wb = self.download_excel(excel_url)
            if wb:
                codici = self.parse_excel_erariali(wb, sezione='erario')
                codici_totali.extend(codici)
        
        print(f"✓ Trovati {len(codici_totali)} codici ERARIO/REGIONI")
        return codici_totali
    
    def scrape_tributi_locali(self) -> List[Dict]:
        """Scarica codici TRIBUTI LOCALI da file Excel"""
        print("\n🏘️  Scaricamento codici TRIBUTI LOCALI...")
        
        # URL diretti ai file Excel tributi locali (aggiornati 28/10/2025)
        # Fonte: https://www.agenziaentrate.gov.it/portale/strumenti/codici-attivita-e-tributo/
        #        f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24/
        #        tabelle-codici-per-tributi-locali
        
        excel_urls = [
            # Tabella per ente (IMU, TASI, TARI)
            f"{self.BASE_URL}/portale/documents/20143/448384/TABELLAENTITRIBUTOxEnte_28_OTTOBRE_2025.xlsx/8fcab388-e011-be93-8dfc-8617662438b7",
            # Tabella per imposta
            f"{self.BASE_URL}/portale/documents/20143/448384/TABELLAENTITRIBUTOxImposta_28_OTTOBRE_2025.xlsx/492d2fa5-d178-6c81-3fe0-d8d8e77795f1"
        ]
        
        codici_totali = []
        
        for excel_url in excel_urls:
            wb = self.download_excel(excel_url)
            if wb:
                # Parse con sezione 'imu' per tributi locali
                codici = self.parse_excel_tributi_locali(wb)
                codici_totali.extend(codici)
        
        print(f"✓ Trovati {len(codici_totali)} codici TRIBUTI LOCALI")
        return codici_totali
    
    def scrape_inps(self) -> List[Dict]:
        """Scarica codici INPS da file Excel"""
        print("\n💼 Scaricamento codici INPS...")
        
        # URL diretti ai file Excel INPS (aggiornati 20/10/2025)
        # Fonte: https://www.agenziaentrate.gov.it/portale/strumenti/codici-attivita-e-tributo/
        #        f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24/
        #        tabelle-codici-inps-e-enti-previdenziali-ed-assicurativi
        
        excel_urls = [
            # Tabella INPS ordinata per codice
            f"{self.BASE_URL}/portale/documents/20143/448375/TAB_INPS_Entprev_20_10_2025_codice.xlsx/e6af9851-64be-0b07-4dd7-baaf3dd9e3b8"
        ]
        
        codici_totali = []
        
        for excel_url in excel_urls:
            wb = self.download_excel(excel_url)
            if wb:
                codici = self.parse_excel_erariali(wb, sezione='inps')
                codici_totali.extend(codici)
        
        print(f"✓ Trovati {len(codici_totali)} codici INPS/ENTI")
        return codici_totali
    
    def scrape_all(self) -> List[Dict]:
        """Scarica tutti i codici tributo"""
        print("=" * 60)
        print("🚀 SCRAPER CODICI TRIBUTO F24 - AGENZIA DELLE ENTRATE (EXCEL)")
        print("=" * 60)
        
        # Scrape tutte le sezioni
        codici_erario = self.scrape_erario_regioni()
        codici_locali = self.scrape_tributi_locali()
        codici_inps = self.scrape_inps()
        
        # Combina tutti
        self.codici = codici_erario + codici_locali + codici_inps
        
        # Statistiche
        print(f"\n📊 TOTALE CODICI SCARICATI: {len(self.codici)}")
        
        if self.codici:
            # Conta per sezione
            sezioni_count = {}
            for c in self.codici:
                sez = c['sezione']
                sezioni_count[sez] = sezioni_count.get(sez, 0) + 1
            
            print("\n📈 Statistiche per sezione:")
            for sez, count in sorted(sezioni_count.items()):
                print(f"  - {sez.upper()}: {count} codici")
        
        return self.codici
    
    def export_csv(self, filepath: str = "scripts/codici_tributo_ade.csv"):
        """Esporta codici in CSV"""
        if not self.codici:
            print("⚠️  Nessun codice da esportare")
            return
        
        try:
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=self.codici[0].keys())
                writer.writeheader()
                writer.writerows(self.codici)
            
            print(f"✓ CSV esportato: {filepath}")
        except Exception as e:
            print(f"❌ Errore export CSV: {e}")
    
    def export_json(self, filepath: str = "scripts/codici_tributo_ade.json"):
        """Esporta codici in JSON"""
        if not self.codici:
            print("⚠️  Nessun codice da esportare")
            return
        
        try:
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            
            output = {
                'metadata': {
                    'fonte': 'Agenzia delle Entrate',
                    'url_base': self.BASE_URL,
                    'data_scaricamento': datetime.now().isoformat(),
                    'totale_codici': len(self.codici)
                },
                'codici': self.codici
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(output, f, ensure_ascii=False, indent=2)
            
            print(f"✓ JSON esportato: {filepath}")
        except Exception as e:
            print(f"❌ Errore export JSON: {e}")
    
    def update_database(self):
        """Aggiorna database Django"""
        if not self.codici:
            print("⚠️  Nessun codice da importare")
            return
        
        try:
            # Setup Django
            import django
            sys.path.append(os.path.dirname(os.path.dirname(__file__)))
            os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mygest.settings')
            django.setup()
            
            from scadenze.models import CodiceTributoF24
            
            print(f"\n💾 Aggiornamento database...")
            
            creati = 0
            aggiornati = 0
            
            for codice_dict in self.codici:
                obj, created = CodiceTributoF24.objects.update_or_create(
                    codice=codice_dict['codice'],
                    sezione=codice_dict['sezione'],
                    defaults={
                        'descrizione': codice_dict['descrizione'],
                        'causale': codice_dict['causale'],
                        'periodicita': codice_dict.get('periodicita', 'Variabile'),
                        'attivo': codice_dict['attivo'],
                        'fonte': codice_dict['fonte']
                    }
                )
                
                if created:
                    creati += 1
                else:
                    aggiornati += 1
            
            print(f"✓ Database aggiornato:")
            print(f"  - Creati: {creati} nuovi codici")
            print(f"  - Aggiornati: {aggiornati} codici esistenti")
            print(f"  - Totale nel DB: {CodiceTributoF24.objects.count()}")
            
        except Exception as e:
            print(f"❌ Errore aggiornamento database: {e}")
            import traceback
            traceback.print_exc()


def main():
    parser = argparse.ArgumentParser(description='Scraper codici tributo F24 da file Excel ADE')
    parser.add_argument('--update-db', action='store_true', help='Aggiorna database Django')
    parser.add_argument('--export-csv', action='store_true', help='Esporta in CSV')
    parser.add_argument('--export-json', action='store_true', help='Esporta in JSON')
    parser.add_argument('--verbose', action='store_true', help='Output dettagliato')
    
    args = parser.parse_args()
    
    # Default: export CSV e JSON se nessuna opzione specificata
    if not any([args.update_db, args.export_csv, args.export_json]):
        args.export_csv = True
        args.export_json = True
    
    # Crea scraper
    scraper = ScraperADEExcel(verbose=args.verbose)
    
    # Scrape tutti i codici
    codici = scraper.scrape_all()
    
    # Export/Update
    if args.export_csv:
        scraper.export_csv()
    
    if args.export_json:
        scraper.export_json()
    
    if args.update_db:
        scraper.update_database()
    
    print(f"\n✅ Operazione completata!")
    print(f"📅 Data: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == '__main__':
    main()
