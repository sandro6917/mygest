#!/usr/bin/env python3
"""
Parser locale per file Excel codici tributo F24 - Agenzia delle Entrate

Legge file Excel scaricati manualmente dalla directory scripts/excel_ade/

File da scaricare da:
https://www.agenziaentrate.gov.it/portale/strumenti/codici-attivita-e-tributo/
f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24

File necessari:
1. TABELLA_EREL_*.xlsx (Codici Erariali e Regionali)
2. TAB_INPS_*.xlsx (Codici INPS)
3. TABELLAENTITRIBUTOxImposta_*.xlsx (Tributi Locali - opzionale, molto grande)

Usage:
    python parse_codici_tributo_excel.py [--update-db] [--export-csv] [--export-json] [--verbose]
"""

import openpyxl
import xlrd  # Per file .xls (vecchio formato)
import re
import json
import csv
from datetime import datetime
from pathlib import Path
import sys
import os
from typing import List, Dict, Optional
import argparse


class ParserCodiciTributo:
    """Parser per file Excel locali con codici tributo F24"""
    
    # Directory con file Excel
    EXCEL_DIR = Path(__file__).parent / "excel_ade"
    
    def __init__(self, verbose=False):
        self.verbose = verbose
        self.codici = []
        
    def log(self, msg):
        """Log con timestamp se verbose"""
        if self.verbose:
            timestamp = datetime.now().strftime("%H:%M:%S")
            print(f"[{timestamp}] {msg}")
    
    def trova_file_excel(self, pattern: str) -> Optional[Path]:
        """Trova file Excel nella directory che matcha il pattern"""
        if not self.EXCEL_DIR.exists():
            print(f"❌ Directory {self.EXCEL_DIR} non trovata!")
            print(f"   Crea la directory e scarica i file Excel da:")
            print(f"   https://www.agenziaentrate.gov.it/portale/strumenti/codici-attivita-e-tributo/")
            return None
        
        # Cerca file che matchano il pattern
        files = list(self.EXCEL_DIR.glob(pattern))
        
        if not files:
            self.log(f"⚠️  Nessun file trovato per pattern: {pattern}")
            return None
        
        # Prendi il più recente se ci sono più file
        file_più_recente = max(files, key=lambda f: f.stat().st_mtime)
        self.log(f"Trovato: {file_più_recente.name}")
        return file_più_recente
    
    def load_excel(self, filepath: Path) -> Optional[openpyxl.Workbook]:
        """Carica file Excel (.xlsx o .xls)"""
        try:
            self.log(f"Caricamento: {filepath.name}")
            
            # Gestisci .xls (vecchio formato)
            if filepath.suffix.lower() == '.xls':
                self.log(f"Rilevato formato .xls (vecchio)")
                # xlrd restituisce un Workbook diverso, gestiamolo separatamente
                return filepath  # Ritorna il path, lo gestiremo dopo
            else:
                # .xlsx (nuovo formato)
                wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
                return wb
        except Exception as e:
            print(f"❌ Errore caricamento {filepath.name}: {e}")
            return None
    
    def parse_excel_erariali(self, filepath: Path) -> List[Dict]:
        """Parse file Excel codici erariali/regionali"""
        codici = []
        
        wb = self.load_excel(filepath)
        if not wb:
            return []
        
        try:
            sheet = wb.active
            self.log(f"Foglio: {sheet.title}, righe: {sheet.max_row}")
            
            # Trova header row (cerca "codice")
            header_row = None
            for row_idx in range(1, min(5, sheet.max_row + 1)):
                row_values = [str(cell.value).lower() if cell.value else "" for cell in sheet[row_idx]]
                if any('codice' in val for val in row_values):
                    header_row = row_idx
                    break
            
            if not header_row:
                header_row = 1
            
            # Leggi header
            headers = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet[header_row]]
            self.log(f"Header: {headers[:10]}...")
            
            # Trova colonne
            idx_codice = self._find_column(headers, ['tributo', 'codice tributo', 'codice', 'cod.'])
            idx_tipo = self._find_column(headers, ['tipo tributo', 'tipo'])
            idx_desc = self._find_column(headers, ['descrizione', 'causale'])
            
            if idx_codice is None:
                print(f"❌ Colonna 'codice' non trovata")
                return []
            
            if idx_desc is None:
                print(f"❌ Colonna 'descrizione' non trovata")
                return []
            
            self.log(f"Colonne: codice={idx_codice}, tipo={idx_tipo}, desc={idx_desc}")
            
            # Parse righe (con limite per velocità)
            max_rows = min(sheet.max_row + 1, 10000)  # Limita a 10k righe
            
            for row_idx in range(header_row + 1, max_rows):
                if self.verbose and row_idx % 500 == 0:
                    self.log(f"Riga {row_idx}/{max_rows}...")
                
                row = sheet[row_idx]
                
                # Codice
                codice = row[idx_codice].value if idx_codice < len(row) else None
                if not codice:
                    continue
                
                codice = str(codice).strip()
                
                # Skip intestazioni/righe vuote
                if len(codice) < 2 or codice.lower() in ['codice', 'cod.', 'none', 'tributo']:
                    continue
                
                # Tipo tributo (Erario, Regioni, etc.)
                tipo_tributo = ""
                if idx_tipo is not None and idx_tipo < len(row):
                    tipo_val = row[idx_tipo].value
                    tipo_tributo = str(tipo_val).strip().lower() if tipo_val else ""
                
                # Descrizione
                descrizione = ""
                if idx_desc is not None and idx_desc < len(row):
                    desc_val = row[idx_desc].value
                    descrizione = str(desc_val).strip() if desc_val else ""
                
                if len(descrizione) < 5:
                    continue
                
                # Identifica sezione (usa tipo_tributo se disponibile)
                if tipo_tributo:
                    if 'erario' in tipo_tributo:
                        sezione = 'erario'
                    elif 'region' in tipo_tributo:
                        sezione = 'regioni'
                    elif 'inps' in tipo_tributo or 'previdenzial' in tipo_tributo:
                        sezione = 'inps'
                    else:
                        sezione = self._identifica_sezione(codice, descrizione)
                else:
                    sezione = self._identifica_sezione(codice, descrizione)
                
                # Obsoleto?
                attivo = not self._e_obsoleto(descrizione)
                
                codice_dict = {
                    'codice': codice,
                    'sezione': sezione,
                    'descrizione': descrizione[:500],  # Limita lunghezza
                    'causale': descrizione[:100],
                    'periodicita': self._estrai_periodicita(descrizione),
                    'attivo': attivo,
                    'data_aggiornamento': datetime.now().date().isoformat()
                }
                
                codici.append(codice_dict)
            
            self.log(f"✓ Estratti {len(codici)} codici")
            
        except Exception as e:
            print(f"❌ Errore parse: {e}")
            import traceback
            if self.verbose:
                traceback.print_exc()
        finally:
            wb.close()
        
        return codici
    
    def parse_excel_inps(self, filepath: Path) -> List[Dict]:
        """Parse file Excel codici INPS (.xls o .xlsx)"""
        
        # Se è .xls usa xlrd
        if filepath.suffix.lower() == '.xls':
            return self._parse_xls_inps(filepath)
        else:
            # Altrimenti usa il metodo standard
            codici = self.parse_excel_erariali(filepath)
            
            # Forza sezione INPS
            for c in codici:
                c['sezione'] = 'inps'
            
            return codici
    
    def _parse_xls_inps(self, filepath: Path) -> List[Dict]:
        """Parse specifico per file .xls INPS"""
        codici = []
        
        try:
            self.log(f"Parse file .xls: {filepath.name}")
            
            # Apri con xlrd
            wb = xlrd.open_workbook(filepath)
            sheet = wb.sheet_by_index(0)
            
            self.log(f"Foglio: {sheet.name}, righe: {sheet.nrows}")
            
            # Trova header
            header_row = None
            for row_idx in range(min(5, sheet.nrows)):
                row_values = [str(sheet.cell_value(row_idx, col_idx)).lower() 
                             for col_idx in range(sheet.ncols)]
                if any('codice' in val or 'causale' in val for val in row_values):
                    header_row = row_idx
                    break
            
            if not header_row:
                header_row = 0
            
            # Leggi header
            headers = [str(sheet.cell_value(header_row, col_idx)).strip().lower() 
                      for col_idx in range(sheet.ncols)]
            self.log(f"Header: {headers[:5]}...")
            
            # Trova colonne
            idx_codice = self._find_column(headers, ['causale', 'codice', 'cod.'])
            idx_desc = self._find_column(headers, ['tipo soggetto', 'descrizione', 'tipo contributo', 'descrizione contributo'])
            
            if idx_codice is None:
                print(f"❌ Colonna 'codice/causale' non trovata in .xls")
                return []
            
            if idx_desc is None:
                print(f"⚠️  Colonna 'descrizione' non trovata in .xls, uso codice come descrizione")
            
            self.log(f"Colonne: codice={idx_codice}, desc={idx_desc}")
            
            # Parse righe
            for row_idx in range(header_row + 1, min(sheet.nrows, 10000)):
                if self.verbose and row_idx % 500 == 0:
                    self.log(f"Riga {row_idx}/{sheet.nrows}...")
                
                # Codice
                codice = str(sheet.cell_value(row_idx, idx_codice)).strip()
                
                # Skip vuote/intestazioni
                if not codice or len(codice) < 2 or codice.lower() in ['codice', 'causale', 'none']:
                    continue
                
                # Tronca codice se troppo lungo (max 10 char per DB)
                codice_originale = codice
                if len(codice) > 10:
                    codice = codice[:10]
                    if self.verbose:
                        self.log(f"Codice troncato: {codice_originale} -> {codice}")
                
                # Descrizione
                descrizione = ""
                if idx_desc is not None:
                    descrizione = str(sheet.cell_value(row_idx, idx_desc)).strip()
                
                # Se non c'è descrizione, usa il codice
                if len(descrizione) < 3:
                    descrizione = f"INPS - Causale {codice}"
                
                codice_dict = {
                    'codice': codice,
                    'sezione': 'inps',
                    'descrizione': descrizione[:500],
                    'causale': descrizione[:100],
                    'periodicita': 'Mensile',  # INPS di solito mensile
                    'attivo': True,
                    'data_aggiornamento': datetime.now().date().isoformat()
                }
                
                codici.append(codice_dict)
            
            self.log(f"✓ Estratti {len(codici)} codici INPS da .xls")
            
        except Exception as e:
            print(f"❌ Errore parse .xls: {e}")
            if self.verbose:
                import traceback
                traceback.print_exc()
        
        return codici
    
    def _find_column(self, headers: List[str], keywords: List[str]) -> Optional[int]:
        """Trova indice colonna"""
        for keyword in keywords:
            for idx, header in enumerate(headers):
                if keyword in header.lower():
                    return idx
        return None
    
    def _identifica_sezione(self, codice: str, descrizione: str) -> str:
        """Identifica sezione del codice"""
        desc_lower = descrizione.lower()
        
        if any(kw in desc_lower for kw in ['imu', 'tasi', 'tari', 'tributi locali']):
            return 'imu'
        if any(kw in desc_lower for kw in ['inps', 'contributi', 'previdenziali']):
            return 'inps'
        if 'inail' in desc_lower:
            return 'inail'
        if any(kw in desc_lower for kw in ['regional', 'irap', 'addizionale regionale']):
            return 'regioni'
        if 'accise' in desc_lower:
            return 'accise'
        
        return 'erario'
    
    def _e_obsoleto(self, descrizione: str) -> bool:
        """Rileva se codice è obsoleto"""
        parole = ['soppresso', 'abolito', 'abrogato', 'cessato', 'dismesso', 'non più']
        return any(p in descrizione.lower() for p in parole)
    
    def _estrai_periodicita(self, descrizione: str) -> str:
        """Estrae periodicità"""
        desc_lower = descrizione.lower()
        
        if 'trimestral' in desc_lower:
            return 'Trimestrale'
        if 'mensil' in desc_lower:
            return 'Mensile'
        if 'annual' in desc_lower:
            return 'Annuale'
        
        return 'Variabile'
    
    def parse_all(self) -> List[Dict]:
        """Parse tutti i file Excel disponibili"""
        print("=" * 60)
        print("📊 PARSER CODICI TRIBUTO F24 - FILE EXCEL LOCALI")
        print("=" * 60)
        print(f"Directory: {self.EXCEL_DIR}")
        print()
        
        # Crea directory se non esiste
        self.EXCEL_DIR.mkdir(exist_ok=True)
        
        # 1. File ERARIALI/REGIONALI (prioritario)
        print("🏛️  File Erariali/Regionali...")
        file_erariali = self.trova_file_excel("TABELLA_EREL_*.xlsx")
        if file_erariali:
            codici_erariali = self.parse_excel_erariali(file_erariali)
            self.codici.extend(codici_erariali)
            print(f"✓ Trovati {len(codici_erariali)} codici ERARIO/REGIONI")
        else:
            print("⚠️  File erariali non trovato (TABELLA_EREL_*.xlsx)")
        
        # 2. File INPS (.xls o .xlsx)
        print("\n💼 File INPS...")
        file_inps = self.trova_file_excel("*INPS*.xls*") or self.trova_file_excel("Causali_INPS*.xls*")
        if file_inps:
            codici_inps = self.parse_excel_inps(file_inps)
            self.codici.extend(codici_inps)
            print(f"✓ Trovati {len(codici_inps)} codici INPS")
        else:
            print("⚠️  File INPS non trovato (Causali_INPS_*.xls)")
        
        # 3. File TRIBUTI LOCALI (opzionale - molto grande)
        print("\n🏘️  File Tributi Locali...")
        file_locali = self.trova_file_excel("TABELLAENTITRIBUTOxImposta_*.xlsx")
        if file_locali:
            print("   ⚠️  File tributi locali trovato ma SKIP (troppo grande)")
            print("   Se necessario, parsarlo separatamente")
        else:
            print("   (opzionale - file non presente)")
        
        # Statistiche
        print(f"\n📊 TOTALE CODICI: {len(self.codici)}")
        
        if self.codici:
            sezioni = {}
            for c in self.codici:
                sez = c['sezione']
                sezioni[sez] = sezioni.get(sez, 0) + 1
            
            print("\n📈 Per sezione:")
            for sez, count in sorted(sezioni.items()):
                print(f"  - {sez.upper()}: {count}")
        
        return self.codici
    
    def export_csv(self, filepath: str = "scripts/codici_tributo_ade.csv"):
        """Esporta in CSV"""
        if not self.codici:
            print("⚠️  Nessun codice da esportare")
            return
        
        try:
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=self.codici[0].keys())
                writer.writeheader()
                writer.writerows(self.codici)
            
            print(f"✓ CSV: {filepath}")
        except Exception as e:
            print(f"❌ Errore CSV: {e}")
    
    def export_json(self, filepath: str = "scripts/codici_tributo_ade.json"):
        """Esporta in JSON"""
        if not self.codici:
            print("⚠️  Nessun codice da esportare")
            return
        
        try:
            Path(filepath).parent.mkdir(parents=True, exist_ok=True)
            
            output = {
                'metadata': {
                    'fonte': 'Agenzia delle Entrate',
                    'data_import': datetime.now().isoformat(),
                    'totale_codici': len(self.codici)
                },
                'codici': self.codici
            }
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(output, f, ensure_ascii=False, indent=2)
            
            print(f"✓ JSON: {filepath}")
        except Exception as e:
            print(f"❌ Errore JSON: {e}")
    
    def update_database(self):
        """Aggiorna database Django"""
        if not self.codici:
            print("⚠️  Nessun codice da importare")
            return
        
        try:
            # Setup Django
            import django
            sys.path.append(os.path.dirname(os.path.dirname(__file__)))
            os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'mygest.settings')
            django.setup()
            
            from scadenze.models import CodiceTributoF24
            from django.db import transaction
            
            print(f"\n💾 Aggiornamento database...")
            
            creati = 0
            aggiornati = 0
            
            # Usa transazione per velocità
            with transaction.atomic():
                for codice_dict in self.codici:
                    if self.verbose and (creati + aggiornati) % 100 == 0:
                        self.log(f"Processati {creati + aggiornati} codici...")
                    
                    # Cerca se esiste già un codice con lo stesso codice (indipendentemente dalla sezione)
                    existing = CodiceTributoF24.objects.filter(codice=codice_dict['codice']).first()
                    
                    if existing:
                        # Aggiorna l'esistente
                        existing.sezione = codice_dict['sezione']
                        existing.descrizione = codice_dict['descrizione']
                        existing.causale = codice_dict['causale']
                        existing.periodicita = codice_dict.get('periodicita', 'Variabile')
                        existing.attivo = codice_dict['attivo']
                        existing.save()
                        aggiornati += 1
                    else:
                        # Crea nuovo
                        CodiceTributoF24.objects.create(
                            codice=codice_dict['codice'],
                            sezione=codice_dict['sezione'],
                            descrizione=codice_dict['descrizione'],
                            causale=codice_dict['causale'],
                            periodicita=codice_dict.get('periodicita', 'Variabile'),
                            attivo=codice_dict['attivo']
                        )
                        creati += 1
            
            print(f"✓ Database aggiornato:")
            print(f"  - Creati: {creati}")
            print(f"  - Aggiornati: {aggiornati}")
            print(f"  - Totale DB: {CodiceTributoF24.objects.count()}")
            
        except Exception as e:
            print(f"❌ Errore database: {e}")
            import traceback
            if self.verbose:
                traceback.print_exc()


def main():
    parser = argparse.ArgumentParser(
        description='Parser locale file Excel codici tributo F24',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Esempi:
  python parse_codici_tributo_excel.py                    # Solo export CSV/JSON
  python parse_codici_tributo_excel.py --update-db        # Aggiorna database
  python parse_codici_tributo_excel.py --verbose          # Output dettagliato
  
File necessari in scripts/excel_ade/:
  - TABELLA_EREL_*.xlsx (Erariali/Regionali)
  - TAB_INPS_*.xlsx (INPS)
  
Scarica da:
  https://www.agenziaentrate.gov.it/portale/strumenti/codici-attivita-e-tributo/
  f24-codici-tributo-per-i-versamenti/tabelle-dei-codici-tributo-e-altri-codici-per-il-modello-f24
        """
    )
    parser.add_argument('--update-db', action='store_true', help='Aggiorna database Django')
    parser.add_argument('--export-csv', action='store_true', help='Esporta in CSV')
    parser.add_argument('--export-json', action='store_true', help='Esporta in JSON')
    parser.add_argument('--verbose', action='store_true', help='Output dettagliato')
    
    args = parser.parse_args()
    
    # Default: export se nessuna opzione
    if not any([args.update_db, args.export_csv, args.export_json]):
        args.export_csv = True
        args.export_json = True
    
    # Parser
    parser_obj = ParserCodiciTributo(verbose=args.verbose)
    
    # Parse
    codici = parser_obj.parse_all()
    
    # Export/Update
    if args.export_csv:
        parser_obj.export_csv()
    
    if args.export_json:
        parser_obj.export_json()
    
    if args.update_db:
        parser_obj.update_database()
    
    print(f"\n✅ Completato!")
    print(f"📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")


if __name__ == '__main__':
    main()
