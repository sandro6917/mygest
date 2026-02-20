from __future__ import annotations

from datetime import timedelta
from itertools import groupby

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST

from mygest.breadcrumbs import set_breadcrumbs

from .forms import (
    ScadenzaBulkOccurrencesForm,
    ScadenzaForm,
    ScadenzaOccorrenzaFormSet,
)
from .models import Scadenza, ScadenzaAlert, ScadenzaOccorrenza
from .services import AlertDispatcher


@login_required
def home(request):
    q = (request.GET.get("q") or "").strip()
    stato = (request.GET.get("stato") or "").strip()
    priorita = (request.GET.get("priorita") or "").strip()

    scadenze_qs = Scadenza.objects.prefetch_related("occorrenze", "pratiche", "assegnatari").order_by("titolo")
    if q:
        scadenze_qs = scadenze_qs.filter(Q(titolo__icontains=q) | Q(descrizione__icontains=q))
    if stato:
        scadenze_qs = scadenze_qs.filter(stato=stato)
    if priorita:
        scadenze_qs = scadenze_qs.filter(priorita=priorita)

    scadenze = scadenze_qs[:200]

    set_breadcrumbs(request, [("Scadenze", None)])
    context = {
        "scadenze": scadenze,
        "q": q,
        "stato": stato,
        "priorita": priorita,
        "stato_choices": Scadenza.Stato.choices,
        "priorita_choices": Scadenza.Priorita.choices,
    }
    return render(request, "scadenze/home.html", context)


@login_required
def detail(request, pk: int):
    scadenza = get_object_or_404(
        Scadenza.objects.prefetch_related(
            "pratiche",
            "fascicoli",
            "documenti",
            "occorrenze__log_eventi",
        ),
        pk=pk,
    )
    occorrenze = scadenza.occorrenze.all().order_by("inizio")

    crumbs = [
        ("Scadenze", reverse("scadenze:home")),
        (scadenza.titolo, None),
    ]
    set_breadcrumbs(request, crumbs)
    return render(
        request,
        "scadenze/detail.html",
        {
            "scadenza": scadenza,
            "occorrenze": occorrenze,
        },
    )


@login_required
def create(request):
    pratica_id = request.GET.get("pratica")
    initial = {}
    if pratica_id:
        initial["pratiche"] = [pratica_id]

    if request.method == "POST":
        form = ScadenzaForm(request.POST)
        formset = ScadenzaOccorrenzaFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            scadenza = form.save(commit=False)
            scadenza.creato_da = request.user
            scadenza.save()
            form.save_m2m()
            formset.instance = scadenza
            formset.save()
            messages.success(request, "Scadenza creata correttamente.")
            return redirect("scadenze:detail", scadenza.pk)
    else:
        form = ScadenzaForm(initial=initial)
        formset = ScadenzaOccorrenzaFormSet()
    set_breadcrumbs(request, [("Scadenze", reverse("scadenze:home")), ("Nuova", None)])
    return render(
        request,
        "scadenze/form.html",
        {
            "form": form,
            "formset": formset,
            "obj": None,
        },
    )


@login_required
def update(request, pk: int):
    scadenza = get_object_or_404(Scadenza, pk=pk)
    if request.method == "POST":
        form = ScadenzaForm(request.POST, instance=scadenza)
        formset = ScadenzaOccorrenzaFormSet(request.POST, instance=scadenza)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            messages.success(request, "Scadenza aggiornata.")
            return redirect("scadenze:detail", scadenza.pk)
    else:
        form = ScadenzaForm(instance=scadenza)
        formset = ScadenzaOccorrenzaFormSet(instance=scadenza)
    set_breadcrumbs(request, [("Scadenze", reverse("scadenze:home")), (scadenza.titolo, reverse("scadenze:detail", args=[scadenza.pk])), ("Modifica", None)])
    return render(
        request,
        "scadenze/form.html",
        {
            "form": form,
            "formset": formset,
            "obj": scadenza,
        },
    )


@login_required
@require_POST
def trigger_alert(request, occorrenza_pk: int):
    occorrenza = get_object_or_404(ScadenzaOccorrenza, pk=occorrenza_pk)
    dispatcher = AlertDispatcher(user=request.user)
    try:
        dispatcher.dispatch(occorrenza)
        messages.success(request, "Alert inviato correttamente.")
    except Exception as exc:  # pragma: no cover
        messages.error(request, f"Errore durante l'invio dell'alert: {exc}")
    return redirect("scadenze:detail", occorrenza.scadenza_id)


@login_required
def bulk_generate(request, pk: int):
    scadenza = get_object_or_404(Scadenza, pk=pk)
    if request.method == "POST":
        form = ScadenzaBulkOccurrencesForm(request.POST)
        if form.is_valid():
            data = form.cleaned_data
            alert_config = data.get("alert_config") or {}
            scadenza.genera_occorrenze_massive(
                start=data["start"],
                end=data.get("end"),
                count=data.get("count"),
                offset_alert_minuti=data["offset_alert_minuti"],
                metodo_alert=data["metodo_alert"],
                alert_config=alert_config,
            )
            messages.success(request, "Occorrenze generate correttamente.")
            return redirect("scadenze:detail", scadenza.pk)
    else:
        form = ScadenzaBulkOccurrencesForm()
    set_breadcrumbs(
        request,
        [
            ("Scadenze", reverse("scadenze:home")),
            (scadenza.titolo, reverse("scadenze:detail", args=[scadenza.pk])),
            ("Generazione massiva", None),
        ],
    )
    return render(
        request,
        "scadenze/bulk_generate.html",
        {
            "form": form,
            "scadenza": scadenza,
        },
    )


@login_required
def scadenziario(request):
    """Vista calendario con tutte le occorrenze raggruppate per giorno."""
    
    # Gestione parametri temporali
    periodo = request.GET.get("periodo", "settimana")
    data_custom = request.GET.get("data")
    stato_filter = request.GET.get("stato", "")
    priorita_filter = request.GET.get("priorita", "")
    
    oggi = timezone.now()
    
    # Calcola range temporale in base al periodo
    if periodo == "oggi":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=1)
        titolo_periodo = "Oggi"
    elif periodo == "settimana":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=7)
        titolo_periodo = "Prossimi 7 giorni"
    elif periodo == "mese":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=30)
        titolo_periodo = "Prossimi 30 giorni"
    elif periodo == "trimestre":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=90)
        titolo_periodo = "Prossimi 90 giorni"
    elif periodo == "custom" and data_custom:
        try:
            data_start = timezone.datetime.fromisoformat(data_custom)
            inizio = timezone.make_aware(data_start) if timezone.is_naive(data_start) else data_start
            fine = inizio + timedelta(days=30)
            titolo_periodo = f"Da {inizio.strftime('%d/%m/%Y')}"
        except (ValueError, TypeError):
            inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
            fine = inizio + timedelta(days=7)
            titolo_periodo = "Prossimi 7 giorni"
    else:
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=7)
        titolo_periodo = "Prossimi 7 giorni"
    
    # Query occorrenze con prefetch degli alert
    occorrenze_qs = ScadenzaOccorrenza.objects.filter(
        inizio__gte=inizio,
        inizio__lt=fine
    ).select_related(
        'scadenza',
        'comunicazione'
    ).prefetch_related(
        'alerts',
        'scadenza__assegnatari',
        'scadenza__pratiche',
    ).order_by('inizio')
    
    # Applica filtri
    if stato_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__stato=stato_filter)
    if priorita_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__priorita=priorita_filter)
    
    occorrenze = list(occorrenze_qs)
    
    # Raggruppa per giorno
    occorrenze_per_giorno = {}
    for data, gruppo in groupby(occorrenze, key=lambda o: timezone.localtime(o.inizio).date()):
        occorrenze_per_giorno[data] = list(gruppo)
    
    # Statistiche
    totale_occorrenze = len(occorrenze)
    totale_alert_pendenti = ScadenzaAlert.objects.filter(
        occorrenza__in=occorrenze,
        stato=ScadenzaAlert.Stato.PENDENTE
    ).count()
    
    set_breadcrumbs(request, [
        ("Scadenze", reverse("scadenze:home")),
        ("Scadenziario", None)
    ])
    
    context = {
        "occorrenze_per_giorno": occorrenze_per_giorno,
        "inizio": inizio,
        "fine": fine,
        "periodo": periodo,
        "titolo_periodo": titolo_periodo,
        "totale_occorrenze": totale_occorrenze,
        "totale_alert_pendenti": totale_alert_pendenti,
        "stato_filter": stato_filter,
        "priorita_filter": priorita_filter,
        "stato_choices": Scadenza.Stato.choices,
        "priorita_choices": Scadenza.Priorita.choices,
    }
    return render(request, "scadenze/scadenziario.html", context)


@login_required
def calendario_visual(request):
    """Vista calendario con FullCalendar.js."""
    set_breadcrumbs(request, [
        ("Scadenze", reverse("scadenze:home")),
        ("Calendario", None)
    ])
    
    context = {
        "stato_choices": Scadenza.Stato.choices,
        "priorita_choices": Scadenza.Priorita.choices,
    }
    return render(request, "scadenze/calendario_visual.html", context)


@login_required
def calendario_events_json(request):
    """API JSON per eventi del calendario (formato FullCalendar)."""
    start = request.GET.get("start")
    end = request.GET.get("end")
    stato_filter = request.GET.get("stato", "")
    priorita_filter = request.GET.get("priorita", "")
    
    if not start or not end:
        return JsonResponse({"error": "Parametri start/end mancanti"}, status=400)
    
    try:
        start_dt = timezone.datetime.fromisoformat(start.replace("Z", "+00:00"))
        end_dt = timezone.datetime.fromisoformat(end.replace("Z", "+00:00"))
    except (ValueError, TypeError):
        return JsonResponse({"error": "Formato data non valido"}, status=400)
    
    # Query occorrenze
    occorrenze_qs = ScadenzaOccorrenza.objects.filter(
        inizio__gte=start_dt,
        inizio__lt=end_dt
    ).select_related('scadenza').prefetch_related('alerts')
    
    # Applica filtri
    if stato_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__stato=stato_filter)
    if priorita_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__priorita=priorita_filter)
    
    # Mappa colori priorità
    color_map = {
        Scadenza.Priorita.CRITICA: "#dc2626",  # rosso
        Scadenza.Priorita.ALTA: "#f59e0b",      # arancione
        Scadenza.Priorita.MEDIA: "#3b82f6",     # blu
        Scadenza.Priorita.BASSA: "#6b7280",     # grigio
    }
    
    # Costruisci eventi per FullCalendar
    events = []
    for occ in occorrenze_qs:
        num_alerts = occ.alerts.count()
        events.append({
            "id": occ.id,
            "title": f"{occ.scadenza.titolo} {'🗓️' if occ.giornaliera else ''}",
            "start": occ.inizio.isoformat(),
            "end": occ.fine.isoformat() if occ.fine else None,
            "allDay": occ.giornaliera,
            "url": reverse("scadenze:detail", args=[occ.scadenza_id]),
            "backgroundColor": color_map.get(occ.scadenza.priorita, "#6b7280"),
            "borderColor": color_map.get(occ.scadenza.priorita, "#6b7280"),
            "extendedProps": {
                "scadenza_id": occ.scadenza_id,
                "scadenza_titolo": occ.scadenza.titolo,
                "descrizione": occ.descrizione or occ.scadenza.descrizione,
                "priorita": occ.scadenza.get_priorita_display(),
                "stato": occ.get_stato_display(),
                "num_alerts": num_alerts,
            }
        })
    
    return JsonResponse(events, safe=False)


@login_required
def export_scadenziario_pdf(request):
    """Export scadenziario in formato PDF."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    
    # Parametri filtro
    periodo = request.GET.get("periodo", "settimana")
    stato_filter = request.GET.get("stato", "")
    priorita_filter = request.GET.get("priorita", "")
    
    oggi = timezone.now()
    
    # Calcola range temporale
    if periodo == "oggi":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=1)
        titolo_periodo = "Oggi"
    elif periodo == "settimana":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=7)
        titolo_periodo = "Prossimi 7 giorni"
    elif periodo == "mese":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=30)
        titolo_periodo = "Prossimi 30 giorni"
    elif periodo == "trimestre":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=90)
        titolo_periodo = "Prossimi 90 giorni"
    else:
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=7)
        titolo_periodo = "Prossimi 7 giorni"
    
    # Query occorrenze
    occorrenze_qs = ScadenzaOccorrenza.objects.filter(
        inizio__gte=inizio,
        inizio__lt=fine
    ).select_related('scadenza').prefetch_related('alerts').order_by('inizio')
    
    if stato_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__stato=stato_filter)
    if priorita_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__priorita=priorita_filter)
    
    occorrenze = list(occorrenze_qs)
    
    # Crea PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Titolo
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#667eea'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph(f"Scadenziario - {titolo_periodo}", title_style))
    elements.append(Spacer(1, 0.3*cm))
    
    # Sottotitolo con info
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=1
    )
    info_text = f"Generato il {oggi.strftime('%d/%m/%Y %H:%M')} | Totale occorrenze: {len(occorrenze)}"
    elements.append(Paragraph(info_text, subtitle_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # Tabella occorrenze
    if occorrenze:
        data = [['Data/Ora', 'Scadenza', 'Descrizione', 'Priorità', 'Stato', 'Alert']]
        
        for occ in occorrenze:
            data_ora = timezone.localtime(occ.inizio).strftime('%d/%m/%Y %H:%M')
            if occ.giornaliera:
                data_ora = timezone.localtime(occ.inizio).strftime('%d/%m/%Y') + " (giorn.)"
            
            titolo = occ.scadenza.titolo[:40] + "..." if len(occ.scadenza.titolo) > 40 else occ.scadenza.titolo
            descrizione = (occ.descrizione or occ.scadenza.descrizione or "")[:50]
            if len(descrizione) > 50:
                descrizione = descrizione[:47] + "..."
            
            priorita_map = {
                'critical': '🔴 Critica',
                'high': '🟠 Alta',
                'medium': '🔵 Media',
                'low': '⚪ Bassa',
            }
            priorita = priorita_map.get(occ.scadenza.priorita, occ.scadenza.get_priorita_display())
            
            stato = occ.get_stato_display()
            num_alerts = occ.alerts.count()
            
            data.append([
                data_ora,
                titolo,
                descrizione,
                priorita,
                stato,
                str(num_alerts)
            ])
        
        table = Table(data, colWidths=[3.5*cm, 5*cm, 6*cm, 2.5*cm, 2.5*cm, 1.5*cm])
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Body
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
        ]))
        
        elements.append(table)
    else:
        elements.append(Paragraph("Nessuna occorrenza trovata per il periodo selezionato.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    
    # Ritorna response
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    filename = f"scadenziario_{oggi.strftime('%Y%m%d_%H%M%S')}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def export_scadenziario_excel(request):
    """Export scadenziario in formato Excel."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Parametri filtro
    periodo = request.GET.get("periodo", "settimana")
    stato_filter = request.GET.get("stato", "")
    priorita_filter = request.GET.get("priorita", "")
    
    oggi = timezone.now()
    
    # Calcola range temporale
    if periodo == "oggi":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=1)
        titolo_periodo = "Oggi"
    elif periodo == "settimana":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=7)
        titolo_periodo = "Prossimi 7 giorni"
    elif periodo == "mese":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=30)
        titolo_periodo = "Prossimi 30 giorni"
    elif periodo == "trimestre":
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=90)
        titolo_periodo = "Prossimi 90 giorni"
    else:
        inizio = oggi.replace(hour=0, minute=0, second=0, microsecond=0)
        fine = inizio + timedelta(days=7)
        titolo_periodo = "Prossimi 7 giorni"
    
    # Query occorrenze
    occorrenze_qs = ScadenzaOccorrenza.objects.filter(
        inizio__gte=inizio,
        inizio__lt=fine
    ).select_related('scadenza').prefetch_related('alerts').order_by('inizio')
    
    if stato_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__stato=stato_filter)
    if priorita_filter:
        occorrenze_qs = occorrenze_qs.filter(scadenza__priorita=priorita_filter)
    
    occorrenze = list(occorrenze_qs)
    
    # Crea workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scadenziario"
    
    # Stili
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Titolo
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f"Scadenziario - {titolo_periodo}"
    title_cell.font = Font(bold=True, size=16, color="667eea")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Sottotitolo
    ws.merge_cells('A2:G2')
    subtitle_cell = ws['A2']
    subtitle_cell.value = f"Generato il {oggi.strftime('%d/%m/%Y %H:%M')} | Totale occorrenze: {len(occorrenze)}"
    subtitle_cell.font = Font(size=10, color="666666")
    subtitle_cell.alignment = Alignment(horizontal='center')
    
    # Header tabella
    headers = ['Data/Ora', 'Scadenza', 'Descrizione', 'Priorità', 'Stato', 'Alert', 'Giornaliera']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Dati
    priorita_map = {
        'critical': 'Critica',
        'high': 'Alta',
        'medium': 'Media',
        'low': 'Bassa',
    }
    
    for idx, occ in enumerate(occorrenze, start=5):
        data_ora = timezone.localtime(occ.inizio).strftime('%d/%m/%Y %H:%M')
        if occ.giornaliera:
            data_ora = timezone.localtime(occ.inizio).strftime('%d/%m/%Y')
        
        ws.cell(row=idx, column=1, value=data_ora).border = border
        ws.cell(row=idx, column=2, value=occ.scadenza.titolo).border = border
        ws.cell(row=idx, column=3, value=occ.descrizione or occ.scadenza.descrizione or "").border = border
        ws.cell(row=idx, column=4, value=priorita_map.get(occ.scadenza.priorita, occ.scadenza.get_priorita_display())).border = border
        ws.cell(row=idx, column=5, value=occ.get_stato_display()).border = border
        ws.cell(row=idx, column=6, value=occ.alerts.count()).border = border
        ws.cell(row=idx, column=7, value="Sì" if occ.giornaliera else "No").border = border
        
        # Colorazione priorità
        if occ.scadenza.priorita == 'critical':
            for col in range(1, 8):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color="fff5f5", end_color="fff5f5", fill_type="solid")
        elif occ.scadenza.priorita == 'high':
            for col in range(1, 8):
                ws.cell(row=idx, column=col).fill = PatternFill(start_color="fffaf0", end_color="fffaf0", fill_type="solid")
    
    # Autosize colonne
    column_widths = [20, 30, 40, 15, 15, 10, 15]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    # Altezza righe
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[4].height = 25
    
    # Salva in BytesIO
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Ritorna response
    response = HttpResponse(
        buffer.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"scadenziario_{oggi.strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


@login_required
def statistiche(request):
    """Vista con statistiche avanzate e grafici."""
    import json
    
    oggi = timezone.now()
    
    # Range temporale: ultimi 90 giorni + prossimi 90 giorni
    inizio = oggi - timedelta(days=90)
    fine = oggi + timedelta(days=90)
    
    # Occorrenze nel periodo
    occorrenze = ScadenzaOccorrenza.objects.filter(
        inizio__gte=inizio,
        inizio__lt=fine
    ).select_related('scadenza').prefetch_related('alerts')
    
    # Statistiche generali
    totale_scadenze = Scadenza.objects.count()
    totale_occorrenze = occorrenze.count()
    occorrenze_completate = occorrenze.filter(stato=ScadenzaOccorrenza.Stato.COMPLETATO).count()
    occorrenze_pendenti = occorrenze.filter(
        stato__in=[ScadenzaOccorrenza.Stato.PENDENTE, ScadenzaOccorrenza.Stato.PROGRAMMATA]
    ).count()
    
    # Distribuzione per priorità
    priorita_stats = {}
    for priorita, label in Scadenza.Priorita.choices:
        count = Scadenza.objects.filter(priorita=priorita).count()
        priorita_stats[label] = count
    
    # Distribuzione per stato
    stato_stats = {}
    for stato, label in ScadenzaOccorrenza.Stato.choices:
        count = occorrenze.filter(stato=stato).count()
        stato_stats[label] = count
    
    # Andamento mensile (ultimi 6 mesi + prossimi 6 mesi)
    from datetime import date
    import calendar
    
    andamento_mensile = []
    for i in range(-6, 7):
        if i == 0:
            continue
        
        if i < 0:
            # Mesi passati
            mese_data = oggi + timedelta(days=30*i)
        else:
            # Mesi futuri
            mese_data = oggi + timedelta(days=30*i)
        
        mese_start = mese_data.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        if mese_start.month == 12:
            mese_end = mese_start.replace(year=mese_start.year + 1, month=1)
        else:
            mese_end = mese_start.replace(month=mese_start.month + 1)
        
        count = occorrenze.filter(
            inizio__gte=mese_start,
            inizio__lt=mese_end
        ).count()
        
        count_completate = occorrenze.filter(
            inizio__gte=mese_start,
            inizio__lt=mese_end,
            stato=ScadenzaOccorrenza.Stato.COMPLETATO
        ).count()
        
        mese_nome = calendar.month_name[mese_start.month][:3]
        anno = mese_start.year
        
        andamento_mensile.append({
            'label': f"{mese_nome} {anno}",
            'totale': count,
            'completate': count_completate,
        })
    
    # Alert: statistiche invii
    totale_alert = ScadenzaAlert.objects.count()
    alert_inviati = ScadenzaAlert.objects.filter(stato=ScadenzaAlert.Stato.INVIATO).count()
    alert_falliti = ScadenzaAlert.objects.filter(stato=ScadenzaAlert.Stato.FALLITO).count()
    alert_pendenti = ScadenzaAlert.objects.filter(stato=ScadenzaAlert.Stato.PENDENTE).count()
    
    # Scadenze più frequenti (per titolo)
    from django.db.models import Count
    scadenze_top = Scadenza.objects.annotate(
        num_occorrenze=Count('occorrenze')
    ).order_by('-num_occorrenze')[:5]
    
    set_breadcrumbs(request, [
        ("Scadenze", reverse("scadenze:home")),
        ("Statistiche", None)
    ])
    
    context = {
        'totale_scadenze': totale_scadenze,
        'totale_occorrenze': totale_occorrenze,
        'occorrenze_completate': occorrenze_completate,
        'occorrenze_pendenti': occorrenze_pendenti,
        'priorita_stats': json.dumps(priorita_stats),
        'stato_stats': json.dumps(stato_stats),
        'andamento_mensile': json.dumps(andamento_mensile),
        'totale_alert': totale_alert,
        'alert_inviati': alert_inviati,
        'alert_falliti': alert_falliti,
        'alert_pendenti': alert_pendenti,
        'scadenze_top': scadenze_top,
    }
    
    return render(request, "scadenze/statistiche.html", context)
